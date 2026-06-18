"""
dalux_excel.py — shared dataset definitions + Excel export for Dalux FM.

The heart of this module is DATASETS: for each entity (work orders, tickets,
estates, buildings, assets) it defines the columns, the lookup context, and how
to stream the records. Both the Excel export and the daily CSV export
(daily_export.py) build on the SAME definitions, so columns never drift.

Public helpers:
    columns_for(which)          -> list of (header, extractor)
    iter_rows(fm, which)        -> yields header row, then flattened data rows
    export_excel(fm, which, path) -> writes a formatted .xlsx, returns row count
    export_<entity>(fm, path)   -> thin wrappers used by the app

Cross-platform: pure Python + openpyxl.
"""

from datetime import datetime, date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from dalux_fm import DaluxFMClient


# --------------------------------------------------------------------------- #
# value helpers
# --------------------------------------------------------------------------- #
def _dt(s):
    """ISO-8601 string -> naive datetime (for a real Excel/sortable date)."""
    if not s:
        return None
    try:
        return datetime.fromisoformat(str(s).replace("Z", "+00:00")).replace(tzinfo=None)
    except (ValueError, TypeError):
        return s


def _names(refs, key, name_map):
    out = []
    for r in refs or []:
        rid = r.get(key)
        out.append(name_map.get(str(rid), rid))
    return "; ".join(str(x) for x in out if x is not None)


def _map(items, id_key):
    return {str(i.get(id_key)): i.get("name") for i in items}


# --------------------------------------------------------------------------- #
# per-entity context (id->name lookups) and columns
# --------------------------------------------------------------------------- #
def _ctx_workorders(fm):
    return {
        "status":   _map(fm.workorder_statuses(), "workOrderStatusId"),
        "priority": _map(fm.workorder_priorities(), "priorityId"),
        "team":     _map(fm.workorder_teams(), "teamId"),
        "building": _map(fm.buildings.list(), "buildingId"),
    }


def _ctx_tickets(fm):
    return {
        "status":   _map(fm.ticket_statuses(), "ticketStatusId"),
        "topic":    _map(fm.ticket_topics(), "topicId"),
        "building": _map(fm.buildings.list(), "buildingId"),
    }


def _ctx_building_only(fm):
    return {"building": _map(fm.buildings.list(), "buildingId")}


def _asset_cls(a, field):
    c = a.get("classification") or {}
    c = c.get("data", c)
    return c.get(field)


DATASETS = {
    "workorders": {
        "title": "Work orders",
        "ctx": _ctx_workorders,
        "records": lambda fm: fm.workorders.iter(),
        "columns": [
            ("Work order ID", lambda w, c: w.get("workOrderId")),
            ("Number",        lambda w, c: w.get("number")),
            ("Name",          lambda w, c: w.get("name")),
            ("Type",          lambda w, c: w.get("type")),
            ("Status",        lambda w, c: c["status"].get(str(w.get("status")), w.get("status"))),
            ("Priority",      lambda w, c: c["priority"].get(str((w.get("priorityRef") or {}).get("priorityId")), "")),
            ("Team",          lambda w, c: c["team"].get(str((w.get("teamRef") or {}).get("teamId")), "")),
            ("Responsible",   lambda w, c: w.get("responsibleUserEmail")),
            ("Buildings",     lambda w, c: _names((w.get("placement") or {}).get("buildingRefs"), "buildingId", c["building"])),
            ("Description",   lambda w, c: w.get("description")),
            ("Start date",    lambda w, c: _dt(w.get("startDate"))),
            ("Deadline",      lambda w, c: _dt(w.get("deadlineDate"))),
            ("Duration (d)",  lambda w, c: w.get("durationDays")),
            ("Created",       lambda w, c: _dt(w.get("createdDate"))),
            ("Last change",   lambda w, c: _dt(w.get("lastChangeDate"))),
            ("Expected cost", lambda w, c: w.get("expectedCost")),
            ("Hours reg.",    lambda w, c: w.get("hoursRegistered")),
            ("Statutory",     lambda w, c: w.get("isStatutory")),
            ("Warranty",      lambda w, c: w.get("hasWarranty")),
        ],
    },
    "tickets": {
        "title": "Tickets",
        "ctx": _ctx_tickets,
        "records": lambda fm: fm.tickets.iter(),
        "columns": [
            ("Ticket ID",   lambda t, c: t.get("ticketId")),
            ("Number",      lambda t, c: t.get("number")),
            ("Topic",       lambda t, c: c["topic"].get(str((t.get("topic") or {}).get("topicId")), "")),
            ("Status",      lambda t, c: c["status"].get(str(t.get("status")), t.get("status"))),
            ("Buildings",   lambda t, c: _names((t.get("placement") or {}).get("buildingRefs"), "buildingId", c["building"])),
            ("Reporter",    lambda t, c: t.get("reporterEmail")),
            ("Description", lambda t, c: t.get("description")),
            ("Created",     lambda t, c: _dt(t.get("createdDate"))),
            ("Last change", lambda t, c: _dt(t.get("lastChangeDate"))),
        ],
    },
    "estates": {
        "title": "Estates",
        "ctx": None,
        "records": lambda fm: fm.estates.iter(),
        "columns": [
            ("Estate ID",   lambda e, c: e.get("estateId")),
            ("Name",        lambda e, c: e.get("name")),
            ("Description", lambda e, c: e.get("description")),
            ("Location ID", lambda e, c: (e.get("locationRef") or {}).get("locationId")),
            ("Last change", lambda e, c: _dt(e.get("lastChangeDate"))),
        ],
    },
    "buildings": {
        "title": "Buildings",
        "ctx": None,
        "records": lambda fm: fm.buildings.iter(),
        "columns": [
            ("Building ID", lambda b, c: b.get("buildingId")),
            ("Name",        lambda b, c: b.get("name")),
            ("Estate ID",   lambda b, c: (b.get("estateRef") or {}).get("estateId")),
            ("Road",        lambda b, c: (b.get("address") or {}).get("road")),
            ("Number",      lambda b, c: (b.get("address") or {}).get("number")),
            ("Zip",         lambda b, c: (b.get("address") or {}).get("zipCode")),
            ("City",        lambda b, c: (b.get("address") or {}).get("city")),
            ("Owned",       lambda b, c: b.get("owned")),
            ("Gross area",  lambda b, c: b.get("grossArea")),
            ("Net area",    lambda b, c: b.get("netArea")),
            ("Last change", lambda b, c: _dt(b.get("lastChangeDate"))),
        ],
    },
    "assets": {
        "title": "Assets",
        "ctx": _ctx_building_only,
        "records": lambda fm: fm.assets.iter(),
        "columns": [
            ("Asset ID",       lambda a, c: a.get("assetId")),
            ("Name",           lambda a, c: a.get("name")),
            ("Classification", lambda a, c: _asset_cls(a, "name")),
            ("Class. code",    lambda a, c: _asset_cls(a, "code")),
            ("Buildings",      lambda a, c: _names((a.get("placement") or {}).get("buildingRefs"), "buildingId", c["building"])),
            ("Description",    lambda a, c: a.get("description")),
            ("Created",        lambda a, c: _dt(a.get("createdDate"))),
            ("Last change",    lambda a, c: _dt(a.get("lastChangeDate"))),
        ],
    },
}


# --------------------------------------------------------------------------- #
# shared accessors (used by Excel export AND daily_export.py)
# --------------------------------------------------------------------------- #
def columns_for(which):
    return DATASETS[which]["columns"]


def iter_rows(fm, which):
    """Yield the header row, then a flattened list per record. Generic enough
    for CSV, Parquet, or anything tabular."""
    d = DATASETS[which]
    ctx = d["ctx"](fm) if d["ctx"] else {}
    columns = d["columns"]
    yield [h for h, _ in columns]
    for rec in d["records"](fm):
        yield [extract(rec, ctx) for _, extract in columns]


# --------------------------------------------------------------------------- #
# Excel export
# --------------------------------------------------------------------------- #
def write_excel(path, sheet_title, columns, records, ctx):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    for col, (title, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    n = 0
    for rec in records:
        ws.append([extract(rec, ctx) for _, extract in columns])
        n += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{n + 1}"
    for col, (title, _) in enumerate(columns, start=1):
        letter = get_column_letter(col)
        longest = len(str(title))
        for r in range(2, n + 2):
            v = ws.cell(row=r, column=col).value
            if isinstance(v, (datetime, date)):
                ws.cell(row=r, column=col).number_format = "yyyy-mm-dd hh:mm"
                longest = max(longest, 16)
            elif v is not None:
                longest = max(longest, len(str(v)))
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), 50)

    wb.save(path)
    return n


def export_excel(fm, which, path):
    d = DATASETS[which]
    ctx = d["ctx"](fm) if d["ctx"] else {}
    return write_excel(path, d["title"], d["columns"], d["records"](fm), ctx)


# thin per-entity wrappers (used by the terminal app)
def export_workorders(fm, path): return export_excel(fm, "workorders", path)
def export_tickets(fm, path):    return export_excel(fm, "tickets", path)
def export_estates(fm, path):    return export_excel(fm, "estates", path)
def export_buildings(fm, path):  return export_excel(fm, "buildings", path)
def export_assets(fm, path):     return export_excel(fm, "assets", path)

# name -> (exporter, default filename stem)
EXPORTERS = {
    "workorders": (export_workorders, "workorders"),
    "tickets":    (export_tickets, "tickets"),
    "estates":    (export_estates, "estates"),
    "buildings":  (export_buildings, "buildings"),
    "assets":     (export_assets, "assets"),
}


if __name__ == "__main__":
    import sys
    fm = DaluxFMClient()
    which = sys.argv[1] if len(sys.argv) > 1 else "workorders"
    if which not in EXPORTERS:
        sys.exit(f"Unknown type '{which}'. Choose: {', '.join(EXPORTERS)}")
    out = sys.argv[2] if len(sys.argv) > 2 else f"{which}_{date.today().isoformat()}.xlsx"
    print(f"Wrote {export_excel(fm, which, out)} {which} to {out}")
